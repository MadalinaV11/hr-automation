import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter

# Load the raw HR data
file_path = "Employees.xlsx"
sheet_name = "Sheet1"
df = pd.read_excel(file_path, sheet_name=sheet_name, dtype={'Start_time': str, 'End_time': str})

# Format Employee Info
df["Employee_Info"] = df["Full_Name"] + "\n" + df["Role"]

# Sort employees by DOJ
df = df.sort_values(by="Local_DOJ")

# Get unique employees
employees = df[["Employee_Info"]].drop_duplicates()

# Get the current month and year
today = datetime.today()
year = today.year
#current month
#month = today.month

#february
month = 2 
start_date = datetime(year, month, 1)  # first day 

# last day (accounts for leap years)
if month == 12:
    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
else:
    end_date = datetime(year, month + 1, 1) - timedelta(days=1)

# Define the date range dynamically
start_date = datetime(year, month, 1)  # First day of the month
end_date = datetime(year, month + 1, 1) - timedelta(days=1)  # Last day of the month

# Create an empty attendance sheet with headers
dates = [day.day for day in pd.date_range(start_date, end_date)]  # Extract day as integer
#attendance_data = pd.DataFrame(columns=[""] + ["Employee_Info"] + dates)  # Empty Column A

# Load the Holidays file
holidays_file = "Holidays.xlsx"
holidays_df = pd.read_excel(holidays_file, sheet_name="Sheet1")

# Identify weekends and public holidays
weekends = {str(day.day) for day in pd.date_range(start_date, end_date) if day.weekday() >= 5}
holidays_df["Date"] = pd.to_datetime(holidays_df["Date"])  # Ensure it's datetime
holidays_df = holidays_df[holidays_df["Date"].dt.month == month]  # Filter holidays for the current month
public_holidays = {str(date.day) for date in holidays_df["Date"]}  # Extract only the day as string

# Fill attendance data
data_rows = []

for _, row in employees.iterrows():
    emp_info = row["Employee_Info"]
    
    # Get employee data from the main DataFrame (df)
    emp_data = df[df["Employee_Info"] == emp_info]
    role = emp_data["Role"].values[0]  # Get role
    #local_doj = pd.to_datetime(emp_data["Local_DOJ"].values[0], errors="coerce")  # Ensure it's a valid datetime

    # Initialize rows for start time, end time, and hours worked
    start_time_row = [""] + [emp_info] + ["" for _ in range(len(dates))]
    end_time_row = [""] + [emp_info] + ["" for _ in range(len(dates))]
    hours_row = [""] + [emp_info] + ["" for _ in range(len(dates))]

    # Get employee working hours from the Employees.xlsx
    start_time = emp_data["Start_time"].values[0] if not emp_data.empty else ""
    end_time = emp_data["End_time"].values[0] if not emp_data.empty else ""
    hours_worked = emp_data["Hours"].values[0] if not emp_data.empty else ""

    # Iterate through each day in the month
    for i, day in enumerate(dates, start=1):  # i starts at 1 for Excel alignment
        day_date = start_date + timedelta(days=i - 1)  # Convert column number to actual date

        # For days after the employee's start date (without Local_DOJ)
        if role == "FEMEIE DE SERVICIU":
            # Only fill if it's Friday (weekday 4)
            if day_date.weekday() == 4:
                start_time_row[i+1] = start_time
                end_time_row[i+1] = end_time
                hours_row[i+1] = hours_worked
        else:
            # For other roles, exclude weekends and public holidays
            if str(day) not in weekends and str(day) not in public_holidays:
                start_time_row[i+1] = start_time
                end_time_row[i+1] = end_time
                hours_row[i+1] = hours_worked

    # Append the rows for this employee
    data_rows.extend([start_time_row, end_time_row, hours_row])

# Convert to DataFrame
attendance_data = pd.DataFrame(data_rows)
attendance_data.columns = [""] + ["Employee_Info"] + dates

# Save to Excel
output_file = "Pontaj.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    attendance_data.to_excel(writer, index=False, startrow=9)  # Headers start at row 10 (index 9)

# Load Excel file for formatting
wb = load_workbook(output_file)
ws = wb.active

# Freeze the first 10 rows
ws.freeze_panes = "A11"  # Freezes rows 1-10, row 11 remains scrollable

# Define borders
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

custom_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       bottom=Side(style='thick'), top=Side(style='none'))

#define colors
#color for numbers
light_red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red fill
#color for weekends and public holidays
grey_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
#color for the days before DOJ
dark_grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

# Apply styles to headers
for row in ws.iter_rows(min_row=10, max_row=10, min_col=2, max_col=ws.max_column):
    for cell in row:
        cell.font = Font(name="Aptos Narrow", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

# Apply merging, fill, and borders for Column A
row = 11  # Start from row 11
employee_number = 1  # Start numbering employees from 1

while row <= ws.max_row:
    ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=1)  # Merge every 3 rows

    for r in range(row, row + 3):  # Iterate through the 3 rows
        cell = ws.cell(row=r, column=1)
        cell.fill = light_red_fill  # Apply light red fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Center text

        # Apply borders
        if r == row:  # First row (thick top, thin sides)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thick"))
            cell.value = employee_number  # Assign number normally
            cell.number_format = '0"." '   # Custom format to display as "1."
        elif r == row + 2:  # Last row (thick bottom, thin sides)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thick"))
        else:  # Middle row (only thin sides)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))

    employee_number += 1
    row += 3  # Move to the next employee block

# style employees
row = 11  # Start from first data row
while row <= ws.max_row:
    ws.merge_cells(start_row=row, start_column=2, end_row=row+2, end_column=2)  # Merge three rows
    cell = ws.cell(row=row, column=2)
    # Apply line break and create the employee info with name and role separated by line break
    cell.value = df.iloc[(row - 11) // 3, df.columns.get_loc("Employee_Info")]
    cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)  # Center text
    cell.font=Font(name="Aptos Narrow", size=11)
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thick"),)
    row += 3  # Move to the next employee's block

# Auto-adjust width for Column A based on longest entry
max_length = max(len(str(cell.value)) for cell in ws["B"] if cell.value)
ws.column_dimensions["B"].width = max_length + 2

# style hours worked
for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
    for cell in row:
        cell.font = Font(name="Aptos Narrow", size=10)  # Set font
        cell.alignment = Alignment(horizontal="right", vertical="center")  # Center all text
        cell.border = thin_border  


# Apply grey fill to weekends & public holidays
for col_num, day_str in enumerate(dates, start=3):
    if str(day_str) in weekends or str(day_str) in public_holidays:
        for row in range(11, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            cell.fill = grey_fill  # Apply grey fill
            cell.value = ""  # Clear text

# Apply dark grey fill to employees who started after the beginning of the current month
# deletes the cell value for those
for row in range(11, ws.max_row + 1, 3):  # Iterate over each employee's data row (skip by 3 since each employee occupies 3 rows)
    emp_info = ws.cell(row=row, column=2).value  # Get employee info (name + role)
    
    # Get the employee's Local_DOJ from the DataFrame
    emp_data = df[df["Employee_Info"] == emp_info] 
    local_doj = pd.to_datetime(emp_data["Local_DOJ"].values[0], errors="coerce")

    # Check if the employee started after the beginning of the current month
    if isinstance(local_doj, pd.Timestamp) and local_doj >= start_date:
        for col_num, day_str in enumerate(dates, start=3):
            day_date = start_date + timedelta(days=col_num - 3)  # Convert column number to actual date

            # Check if the day is before the employee's Local_DOJ
            if isinstance(day_date, datetime) and day_date < local_doj:
                # Apply dark grey to all 3 rows (the start_time_row, end_time_row, hours_row for this employee)
                for row_offset in range(0, 3):  # Employee occupies 3 rows
                    cell = ws.cell(row=row + row_offset, column=col_num)
                    cell.fill = dark_grey_fill  # Apply dark grey fill
                    cell.value = ""  # Empty the cell (before the employee started)


# 8.00 format
for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, (int,float)):  # Only apply to numerical values
            cell.number_format = '0.00'  # Two decimal places
            

wb.save(output_file)
print(f"Formatted file saved as: {output_file}")
