import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import defaultdict
from calendar import month_name



# Insert your data source here
#replace <root_folder> with your file path. 

emp_table = r"<root_folder>\Employees.xlsx" 
PTO_table = r"<root_folder>\PTO.xlsx"
output_file = r"<root_folder>\Pontaj.xlsx"
holidays_file = r"<root_folder>\Holiday_table.xlsx"



# Read Employee table
df = pd.read_excel(emp_table, sheet_name='Sheet1')


# Create employee info column
df["Employee_Info"] = df["Full_Name"] + "\n" + df["Role"]


# sort by DOJ
df = df.sort_values(by="Local_DOJ")


# unique emp
employees = df[['Employee_Info', 'Fusion_ID']].drop_duplicates(subset=['Employee_Info'])


# Current year
today = datetime.today()
year = today.year
#current month
#month = today.month


#Trial with february data
month = 2#TODO: DE SCOS DE AICI CA E PENTRU TEST
start_date = datetime(year, month, 1)  # first day 



# last day (accounts for leap years)
if month == 12:
    end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
else:
    end_date = datetime(year, month + 1, 1) - timedelta(days=1)


# Create an empty attendance sheet with headers as integers
dates = [day.day for day in pd.date_range(start_date, end_date)] 
attendance_data = pd.DataFrame(columns=["Employee_Info"] + dates)



# Get all legal vacation days 
holidays_data = pd.read_excel(holidays_file)
holiday_dates = holidays_data["date"]


# Convert legal vacation days to datetime and filter for the selected month
holiday_dates = [datetime.strptime(date, "%Y/%m/%d") for date in holiday_dates if datetime.strptime(date, "%Y/%m/%d").month == month]


# Extract only the days as strings
public_holidays = {str(date.day) for date in holiday_dates}


# Identify weekends and public holidays
weekends = {str(day.day) for day in pd.date_range(start_date, end_date) if day.weekday() >= 5}


# Define leave abbreviation mappings
leave_abbr_dict = {
    'RO Earned Leave': 'CO',
    'RO Wellness Days': 'ZLP',
    'RO Birthday leave': 'ZLP',
    'RO Vaccination Dose': 'N/A',
    'RO Sick Leave': 'BO',
    'RO Bereavement Leave': 'ZLP',
    'RO Marriage Leave': 'ZLP',
    'RO Relocation Leave': 'ZLP',
    'RO Children Marriage Leave': 'ZLP',
    'RO Childbirth Leave': 'ZLP',
    'RO Pregnancy Leave': 'M',
    'RO Maternity Risk Leave': 'BO',
    'RO Blood Donation Leave': 'ZLP',
    'RO Unpaid Time Off': 'CFS',
    'Ro Adoption/ Childcare leave': 'CCC'
}

# Get PTO table data
pto = pd.read_excel(PTO_table, sheet_name="Sheet1")



# Filter leave PTO table for approved and submitted absences
pto = pto[
    (pto["APPROVAL_STATUS_CD"] == "APPROVED") & 
    (pto["ABSENCE_STATUS_CD"] == "SUBMITTED")
]


# Create a dictionary to store leave days and
# Get data from PTO table to check days off per person

days_off_per_person = defaultdict(lambda: defaultdict(str))
for _, row in pto.iterrows():
    person_number = row["PERSON_NUMBER"]
    absence_type = row["ABSENCE_TYPE"]
    days_off_start_date = row["START_DATE"]
    days_off_end_date = row["END_DATE"]


# verific daca sunt de format date si daca nu le convertesc in date    ///modified by Alex
    if isinstance(days_off_start_date, datetime):
        days_off_start_date = days_off_start_date.date()
    else:
        days_off_start_date = datetime.strptime(days_off_start_date, "%d-%b-%Y").date()


    if isinstance(days_off_end_date, datetime):
        days_off_end_date = days_off_end_date.date()
    else:
        days_off_end_date = datetime.strptime(days_off_end_date, "%d-%b-%Y").date()


# eroare pt abrevieri (daca vor fi greseli in fisierul pto)
    if absence_type not in leave_abbr_dict:
        raise KeyError(f"ABSENCE_TYPE-ul '{absence_type}' pentru persoana '{person_number}' nu exista!")



# aici pun abrevierile
    current_date = days_off_start_date
    while current_date <= days_off_end_date:
        current_date_str = current_date.strftime('%Y-%m-%d')
        days_off_per_person[person_number][current_date_str] = leave_abbr_dict[absence_type]
        current_date += timedelta(days=1)


# ADAUGAREA DE DATE IN EXCEL
print(f"{"\u23F3"} Start adding employee data.\n")

# Fill attendance data
data_rows = []

for _, row in employees.iterrows():
    emp_info = row["Employee_Info"]
    
    # Get employee data from the main DataFrame (df) - name+role
    emp_data = df[df["Employee_Info"] == emp_info]
    role = emp_data["Role"].values[0]  # Get role


    # Create three rows per employee
    # Initialize rows for start time, end time, and hours worked
    start_time_row = [""] + [emp_info] + ["" for _ in range(len(dates))]
    end_time_row = [""] + [emp_info] + ["" for _ in range(len(dates))]
    hours_row = [""] + [emp_info] + ["" for _ in range(len(dates))]


    # Get employee working hours from the Employees.xlsx
    start_time = emp_data["Start_time"].values[0] if not emp_data.empty else ""
    end_time = emp_data["End_time"].values[0] if not emp_data.empty else ""
    hours_worked = emp_data["Hours"].values[0] if not emp_data.empty else ""


    # Fill working hours dynamically based on employee data
    for i, day in enumerate(dates, start=1):  
        day_date = start_date + timedelta(days=i - 1)  # Convert column number to actual date

        #exclude weekends, public holidays, and pto (personal time off)
        if str(day) not in weekends and str(day) not in public_holidays:
            #femeia de serviciu works only fridays 
            if role == "FEMEIE DE SERVICIU": #and str(day) not in weekends and str(day) not in public_holidays
                if day_date.weekday() == 4: #fill only fridays
                    start_time_row[i+1] = start_time
                    end_time_row[i+1] = end_time
                    hours_row[i+1] = hours_worked
            # For other roles
            else:
                day_date_str = day_date.strftime('%Y-%m-%d')
                if row["Fusion_ID"] in days_off_per_person and day_date_str in days_off_per_person[row["Fusion_ID"]]:
                    #it has PTO this day
                    start_time_row[i+1] = ''
                    end_time_row[i+1] = ''
                    hours_row[i+1] = days_off_per_person[row["Fusion_ID"]][day_date_str]
                else:
                    #does not have PTO
                    start_time_row[i+1] = start_time
                    end_time_row[i+1] = end_time
                    hours_row[i+1] = hours_worked
                    

    # Append rows
    data_rows.extend([start_time_row, end_time_row, hours_row])


# Convert to DataFrame
attendance_data = pd.DataFrame(data_rows)
attendance_data.columns = [""] + ["Employee_Info"] + dates


# Save to Excel
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    attendance_data.to_excel(writer, index=False, startrow=9)  # Headers start at row 10 (index 9)

print(f"{"\u2705"} Finished adding data. \n")


# Load Excel file for formatting
wb = load_workbook(output_file)
ws = wb.active






# FORMATARE DE DATE
print(f"{"\u23F3"} Start formatting base table. \n")


#freeze the first 10 rows
ws.freeze_panes = "A11"  # Freezes rows 1-10, row 11 remains scrollable


# Custom font settings for the Employee Info column
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'))


#thicker top and bottom 
custom_border=Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thick'), bottom=Side(style='thick'))


#define colors
#color for numbers
light_red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red fill

#color for weekends and public holidays
grey_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

#color for the days before DOJ
dark_grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

#color for PTO
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Apply styles to headers
for row in ws.iter_rows(min_row=10, max_row=10, min_col=2, max_col=ws.max_column):
    for cell in row:
        cell.font = Font(name="Aptos Narrow", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


# Apply merging, fill, and borders for Column A
row = 11  # Start from row 11
employee_number = 1  # Start numbering employees from 1



while row <= ws.max_row:
    ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=1)  # Merge every 3 rows

    for r in range(row, row + 3):  # Iterate through the 3 rows
        cell = ws.cell(row=r, column=1)
        cell.fill = light_red_fill  # Apply light red fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Center text

        # Apply borders
        if r == row:  # First row (thick top, thin sides)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thick"), bottom=Side(style="thick"))
            cell.value = employee_number  # Assign number normally
            cell.number_format = '0"." '   # Custom format to display as "1."
            cell.font = Font(name="Aptos Narrow", size=10)
        elif r == row + 2:  # Last row (thick bottom, thin sides)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thick"))
        else:  # Middle row (only thin sides)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))

    employee_number += 1
    row += 3  # Move to the next employee block



# style employees
row = 11  # Start from first data row
while row <= ws.max_row:
    ws.merge_cells(start_row=row, start_column=2, end_row=row+2, end_column=2)  # Merge three rows
    cell = ws.cell(row=row, column=2)
    # Apply line break and create the employee info with name and role separated by line break
    cell.value = df.iloc[(row - 11) // 3, df.columns.get_loc("Employee_Info")]
    cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)  # Center text
    cell.font=Font(name="Aptos Narrow", size=11)
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thick"), bottom=Side(style="thick"))
    row += 3  # Move to the next employee's block



# Auto-adjust width for the emp column based on longest entry
max_length = max(len(str(cell.value)) for cell in ws["B"] if cell.value)
ws.column_dimensions["B"].width = max_length + 2


# style hours worked
for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
    for cell in row:
        cell.font = Font(name="Aptos Narrow", size=10)  # Set font
        cell.alignment = Alignment(horizontal="right", vertical="center")  # Center all text
        cell.border = thin_border  


# grey fill to weekends & public holidays
for col_num, day_str in enumerate(dates, start=3):
    if str(day_str) in weekends or str(day_str) in public_holidays:
        for row in range(11, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            cell.fill = grey_fill  # Apply grey fill
            cell.value = ""  # Clear text


# dark grey fill to employees who started after the beginning of the current month
# + deletes cell value
for row in range(11, ws.max_row + 1, 3):  # Iterate over each employee's data row (skip by 3 since each employee occupies 3 rows)
    emp_info = ws.cell(row=row, column=2).value  # Get employee info (name + role)
    
    # Get the employee's Local_DOJ from the DataFrame
    emp_data = df[df["Employee_Info"] == emp_info] 
    local_doj = pd.to_datetime(emp_data["Local_DOJ"].values[0], errors="coerce")

    # Check if the employee started after the beginning of the current month
    if isinstance(local_doj, pd.Timestamp) and local_doj >= start_date:
        for col_num, day_str in enumerate(dates, start=3):
            day_date = start_date + timedelta(days=col_num - 3)  # Convert column number to actual date

            # Check if the day is before the employee's Local_DOJ
            if isinstance(day_date, datetime) and day_date < local_doj:
                # Apply dark grey to all 3 rows (the start_time_row, end_time_row, hours_row for this employee)
                for row_offset in range(0, 3):  # Employee occupies 3 rows
                    cell = ws.cell(row=row + row_offset, column=col_num)
                    cell.fill = dark_grey_fill  # Apply dark grey fill
                    cell.value = ""  # Empty the cell (before the employee started)

    person_number = emp_data["Fusion_ID"].values[0]
    for col_num, day_str in enumerate(dates, start=3):
        day_date = start_date + timedelta(days=col_num - 3)  # Convert column number to actual date

        if isinstance(day_date, datetime):
            if str(day_str) in weekends or str(day_str) in public_holidays:
                continue

            day_date_str = day_date.strftime('%Y-%m-%d')
            if person_number in days_off_per_person and day_date_str in days_off_per_person[person_number]:
                for row_offset in range(0, 3):  # Employee occupies 3 rows
                    cell = ws.cell(row=row + row_offset, column=col_num)
                    cell.fill = yellow_fill  # Apply yellow fill


# 8.00 format
for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=3, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, (int,float)):  # Only apply to numerical values
            cell.number_format = '0.00'  # Two decimal places


ws["N5"].value = "Foaie colectiva de prezenta (pontaj)"
ws["N5"].font = Font(name="Aptos Narrow", size=14, underline="single", bold=True)
ws["N5"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("N5:R5")
ws["D2"].value = "BRILLIO ROMANIA SRL"
ws.merge_cells("D2:F2")
ws["D3"].value = "RO 35553895, J05/223/2016"
ws.merge_cells("D3:F3")
ws["D4"].value = "Oradea, Str Albacului, nr. 12A"
ws.merge_cells("D3:F3")


# Calculează totalul zilelor din lună
total_days = (end_date - start_date).days + 1


# Identifică zilele libere care NU sunt în weekend
public_holidays_weekdays = {str(date.day) for date in holiday_dates if date.weekday() < 5}


# Calculează zilele lucrătoare
working_days = total_days - len(weekends) - len(public_holidays_weekdays)


month_name_ro = {
    "January": "Ianuarie", "February": "Februarie", "March": "Martie",
    "April": "Aprilie", "May": "Mai", "June": "Iunie",
    "July": "Iulie", "August": "August", "September": "Septembrie",
    "October": "Octombrie", "November": "Noiembrie", "December": "Decembrie"
}
current_month_name = month_name_ro[month_name[month]]


# Creează textul pentru celulă
luna_text = f"Luna: {current_month_name} - {working_days} zile lucrătoare"

ws["K7"] = luna_text
ws["K7"].font = Font(size=12, bold=True)
ws.merge_cells("K7:O7")


ws["U7"] = f"Anul: {year}"
ws["U7"].font = Font(size=12, bold=True)
ws.merge_cells("U7:V7")


print(f"{"\u2705"} Format base table done.\n")
# print(f"Format done. File saved in: {output_file}")






# PART 3: Add the calculation table
print(f"{"\u23F3"} Start adding calculation table\n")


# Define the headers for row 9 - max col. nr +5 
row_9_headers = [
    "Total ore lucrate", "din care:", "", "Total ore lucrate", "Total zile lucrate"
]

# Find the last column by checking the maximum column index
last_column = ws.max_column


# Set row 9 headers
for i, header in enumerate(row_9_headers, start=1):
    ws.cell(row=9, column=last_column + i, value=header)



# Define the headers for row 11 - max col. nr + 15
row_10_headers = [
    "", "ore supl", "ore de noapte", "", "", "Co", "Bo", "Bp", "Am", "M", 
    "ZLP", "O", "N", "Pm", "Prb", "Cs"
]



# Set the row 11 headers
for i, header in enumerate(row_10_headers, start=1):
    ws.cell(row=10, column=last_column + i, value=header)



# Add formula for each column row: i+3
start_index = 1
end_index = 17
max_Rows = len(ws['A'])
last_col_name = get_column_letter(last_column)



# Add the number of columns based on range (start_index, end_index) 
for i in range(start_index, end_index):
    current_col = last_column+i # Get new column letter
    current_col_name = get_column_letter(current_col)
    start_row = 13


    # Define dynamic column references. Use table last_col in order to have an absolute column name
    a = get_column_letter(last_column + 1)
    b = get_column_letter(last_column + 2)
    c = get_column_letter(last_column + 3)
    d = get_column_letter(last_column + 4)

    # Iterate over each row in table, and for each cell in the range mentioned in iter_rows, add formula based on his table position 
    for row in ws.iter_rows(min_row=start_row, max_row=max_Rows, min_col=current_col, max_col= current_col):

        formulas = {                                                 # Define formulas dynamically
            "1": f"=SUM(C{start_row}:${last_col_name}{start_row})",
            "2": 0,
            "3": 0,
            "4": f"={a}{start_row}+{b}{start_row}+{c}{start_row}",
            "5": f"={d}{start_row}/8",
            "6": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "7": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "8": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "9": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "10": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "11": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "12": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "13": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "14": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "15": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8",
            "16": f"=COUNTIF($C{start_row}:${last_col_name}{start_row},${current_col_name}$10)*8"
        }

        # Get the formula according to the column nr.
        formula = formulas[str(i)]
        

        # Check if start row smaller than max rows, 
        # Insert the formula into the correct column and row 
        # Move down 3 rows for the next insertion
        if start_row <= max_Rows:
            ws[f"{current_col_name}{start_row}"] = formula
            start_row += 3

# Format headers, subheaders and cells based 
for row in ws.iter_rows(min_row=9, max_row=max_Rows, min_col=last_column+1): # format cells
    if row[0].row < 10: #format row 10 and upward
        for cell in row:
            if cell.value in ["Total ore lucrate", "Total zile lucrate"]:

                cell.font = Font(name="Arial", size=8, bold=True)
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            else:
                for cell in row:
                    cell.font = Font(name="Arial", size=8)
                    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    elif row[0].row == 10: #format row 11 
        for cell in row:
            cell.font = Font(name="Arial", size=7, bold=True)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    else:
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


# format first col in calculation table
for row in ws.iter_rows(min_row=13, max_row=max_Rows, min_col=last_column+1, max_col=last_column+1):
    for cell in row:
        cell.font = Font(name="Arial", size=10, bold=True) 


# format 'Total ore lucrate' col
def add_bottom_border():
    start_r = start_row
    for row in ws.iter_rows(min_row=start_r, max_row=max_Rows, min_col=last_column+4, max_col=last_column+4):
        for cell in row:
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thick"))
        start_r +=3

add_bottom_border()

# format CO sums  
for row in ws.iter_rows(min_row=13, max_row=max_Rows, min_col=last_column+6):
    for cell in row:
        cell.font = Font(name="Arial", size=10, bold=True)



# Save the modified workbook
wb.save(output_file)
print(f"{"\u2705"} Modification done. File saved")
